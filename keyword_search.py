"""
Keyword Search Tool for Candidate Excel Data
=============================================
Searches candidate records in the Trend Ecosystem Excel file for user-specified
keywords. Matches are found across key text fields (Resume Headline, Key Skills,
Company, Designation, etc.) and results are displayed with candidate details
plus their Naukri profile links.

Usage:
    python keyword_search.py
    Then follow the interactive prompts to enter keywords.
"""

import os
import sys
import io
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Fix Windows console encoding — force UTF-8 output
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# ── Configuration ────────────────────────────────────────────────────────────
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "Copy of Trend Ecosystem.xlsx")

# Columns to search for keywords (0-indexed internally, 1-indexed in Excel)
SEARCHABLE_COLUMNS = {
    "Job Title": 1,
    "Name": 3,
    "Current Location": 6,
    "Preferred Locations": 7,
    "Total Experience": 8,
    "Company": 9,
    "Designation": 10,
    "Resume Headline": 13,
    "UG Degree": 14,
    "PG Specialization": 16,
    "Key Skills": 29,
}

# Columns to display in results
DISPLAY_COLUMNS = {
    "Name": 3,
    "Email": 4,
    "Phone": 5,
    "Location": 6,
    "Experience": 8,
    "Company": 9,
    "Designation": 10,
    "Salary": 11,
    "Notice Period": 12,
    "Resume Headline": 13,
    "Key Skills": 29,
    "Trend Micro Cloud Exp": 30,
    "Vision One Exp": 31,
    "Apex One Exp": 32,
    "Expected CTC": 33,
    "Current CTC": 36,
    "Pipeline Stage": 23,
}

PROFILE_LINK_COLUMN = 37  # "Candidate profile" column with hyperlinks


# ── Helpers ──────────────────────────────────────────────────────────────────
def load_workbook_data(filepath, sheet_name=None):
    """Load the workbook and return the target worksheet."""
    if not os.path.exists(filepath):
        print(f"\n[ERROR] File not found: {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    return wb, ws


def get_data_row_count(ws):
    """Count rows with actual data (non-empty Name column)."""
    count = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=3).value:
            count += 1
        else:
            break
    return count


def cell_to_str(value):
    """Safely convert a cell value to a lowercase string for matching."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).lower().strip()


def get_profile_link(ws, row):
    """Extract the hyperlink URL from the profile column."""
    cell = ws.cell(row=row, column=PROFILE_LINK_COLUMN)
    if cell.hyperlink:
        return cell.hyperlink.target
    return cell.value if cell.value else "N/A"


def search_candidates(ws, keywords, total_rows, match_mode="any"):
    """
    Search candidates for keywords across searchable columns.

    Args:
        ws: The worksheet to search.
        keywords: List of keyword strings (lowercase).
        total_rows: Number of data rows.
        match_mode: "any" = match any keyword, "all" = match all keywords.

    Returns:
        List of dicts with candidate info and matched keywords.
    """
    results = []

    for row in range(2, 2 + total_rows):
        # Build a combined text blob from all searchable columns
        searchable_texts = {}
        combined_text = ""
        for col_name, col_idx in SEARCHABLE_COLUMNS.items():
            text = cell_to_str(ws.cell(row=row, column=col_idx).value)
            searchable_texts[col_name] = text
            combined_text += " " + text

        # Check keyword matches
        matched_keywords = []
        matched_in_fields = []
        for kw in keywords:
            if kw in combined_text:
                matched_keywords.append(kw)
                # Find which fields matched
                fields = [name for name, text in searchable_texts.items() if kw in text]
                matched_in_fields.extend(fields)

        # Apply match mode
        if match_mode == "all" and len(matched_keywords) < len(keywords):
            continue
        if match_mode == "any" and len(matched_keywords) == 0:
            continue

        # Collect display data
        candidate = {"row": row, "matched_keywords": matched_keywords,
                      "matched_fields": list(set(matched_in_fields))}
        for col_name, col_idx in DISPLAY_COLUMNS.items():
            val = ws.cell(row=row, column=col_idx).value
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            candidate[col_name] = val if val else "N/A"

        candidate["Profile Link"] = get_profile_link(ws, row)
        results.append(candidate)

    return results


def display_results(results, keywords):
    """Pretty-print the search results."""
    if not results:
        print("\n" + "=" * 60)
        print("  [!] No candidates found matching your keywords.")
        print("=" * 60)
        return

    print("\n" + "=" * 70)
    print(f"  Found {len(results)} candidate(s) matching: {', '.join(keywords)}")
    print("=" * 70)

    for i, c in enumerate(results, 1):
        print(f"\n{'-' * 70}")
        print(f"  #{i}  {c['Name']}")
        print(f"{'-' * 70}")
        print(f"  Email           : {c['Email']}")
        print(f"  Phone           : {c['Phone']}")
        print(f"  Location        : {c['Location']}")
        print(f"  Experience      : {c['Experience']}")
        print(f"  Company         : {c['Company']}")
        print(f"  Designation     : {c['Designation']}")
        print(f"  Current Salary  : {c['Salary']}")
        print(f"  Current CTC     : {c['Current CTC']}")
        print(f"  Expected CTC    : {c['Expected CTC']}")
        print(f"  Notice Period   : {c['Notice Period']}")
        print(f"  Pipeline Stage  : {c['Pipeline Stage']}")
        print(f"  Cloud Exp       : {c['Trend Micro Cloud Exp']}")
        print(f"  Vision One Exp  : {c['Vision One Exp']}")
        print(f"  Apex One Exp    : {c['Apex One Exp']}")
        print(f"  Resume Headline : {c['Resume Headline']}")
        print(f"  Key Skills      : {c['Key Skills']}")
        print(f"  Matched Keywords: {', '.join(c['matched_keywords'])}")
        print(f"  Matched In      : {', '.join(c['matched_fields'])}")
        print(f"  Profile Link    : {c['Profile Link']}")

    print(f"\n{'=' * 70}")


def export_results(results, keywords, output_file="search_results.xlsx"):
    """Export matched candidates to a new Excel file."""
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Search Results"

    # Write header
    headers = ["#", "Name", "Email", "Phone", "Location", "Experience",
               "Company", "Designation", "Current Salary", "Current CTC",
               "Expected CTC", "Notice Period", "Pipeline Stage",
               "Cloud Exp", "Vision One Exp", "Apex One Exp",
               "Resume Headline", "Key Skills",
               "Matched Keywords", "Matched Fields", "Profile Link"]
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Write data
    for i, c in enumerate(results, 1):
        row_data = [
            i, c["Name"], c["Email"], c["Phone"], c["Location"],
            c["Experience"], c["Company"], c["Designation"],
            c["Salary"], c["Current CTC"], c["Expected CTC"],
            c["Notice Period"], c["Pipeline Stage"],
            c["Trend Micro Cloud Exp"], c["Vision One Exp"], c["Apex One Exp"],
            c["Resume Headline"], c["Key Skills"],
            ", ".join(c["matched_keywords"]),
            ", ".join(c["matched_fields"]),
            c["Profile Link"],
        ]
        for col, val in enumerate(row_data, 1):
            ws_out.cell(row=i + 1, column=col, value=val)

    # Auto-adjust column widths (approximate)
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws_out.cell(row=r, column=col).value or ""))
            for r in range(1, len(results) + 2)
        )
        ws_out.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_file)
    wb_out.save(output_path)
    print(f"\n  Results exported to: {output_path}")


# ── Main Interactive Loop ────────────────────────────────────────────────────
def main():
    print("\n" + "=" * 70)
    print("  Candidate Keyword Search Tool")
    print("  File: Copy of Trend Ecosystem.xlsx")
    print("=" * 70)

    wb, ws = load_workbook_data(EXCEL_FILE)
    sheets = wb.sheetnames
    total_rows = get_data_row_count(ws)

    # Sheet selection
    if len(sheets) > 1:
        print(f"\n  Available sheets:")
        for i, name in enumerate(sheets, 1):
            print(f"    {i}. {name}")
        choice = input(f"\n  Select sheet [1-{len(sheets)}] (default: 1): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(sheets):
            ws = wb[sheets[int(choice) - 1]]
            total_rows = get_data_row_count(ws)

    print(f"\n  Loaded {total_rows} candidates from sheet: '{ws.title}'")
    print(f"\n  Searchable fields: {', '.join(SEARCHABLE_COLUMNS.keys())}")

    while True:
        print(f"\n{'-' * 70}")
        raw = input("  Enter keywords (comma-separated, or 'quit' to exit): ").strip()

        if raw.lower() in ("quit", "exit", "q"):
            print("\n  Goodbye!\n")
            break

        if not raw:
            print("  Please enter at least one keyword.")
            continue

        keywords = [kw.strip().lower() for kw in raw.split(",") if kw.strip()]

        # Match mode
        mode_input = input("  Match mode - [A]ny keyword (default) or A[L]l keywords? ").strip().lower()
        match_mode = "all" if mode_input in ("l", "all") else "any"

        print(f"\n  Searching for: {keywords}  (mode: {match_mode})")

        results = search_candidates(ws, keywords, total_rows, match_mode)
        display_results(results, keywords)

        if results:
            export_choice = input("\n  Export results to Excel? [y/N]: ").strip().lower()
            if export_choice in ("y", "yes"):
                filename = f"search_results_{'_'.join(keywords)}.xlsx"
                # Sanitize filename
                filename = "".join(c if c.isalnum() or c in "._-" else "_" for c in filename)
                export_results(results, keywords, filename)


if __name__ == "__main__":
    main()
