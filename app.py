"""
Streamlit Candidate Keyword Search App
=======================================
Dark-themed web interface for searching candidate records from Excel files.
- Multi-file upload with drag & drop
- Multi-keyword tagging
- ATS score calculation from resume data
- Dynamic column visibility
- Experience & CTC range filters
- Excel export with profile links
"""

import os
import io
import re
import math
import html as html_mod
import time
import threading
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
import requests

# ── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ATS Checker",
    page_icon="✅",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Constants ────────────────────────────────────────────────────────────────
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "Copy of Trend Ecosystem.xlsx")

# Short display names for long column headers
HEADER_SHORT_NAMES = {
    "Ans(Key Skills)": "Key Skills",
    "Ans(How many years of experience do you have in Trend Micro Cloud?)": "Trend Micro Cloud Exp",
    "Ans(How many years of experience do you have in Trend Micro Vision One?)": "Vision One Exp",
    "Ans(How many years of experience do you have in Trend Micro Apex One?)": "Apex One Exp",
    "Ans(What is your expected CTC in Lacs per annum?)": "Expected CTC (LPA)",
    "Ans(What is your notice period?)": "Notice Period (Ans)",
    "Ans(Are you currently residing in Mumbai or willing to relocate to Mumbai?)": "Mumbai Relocation",
    "Ans(What is your current CTC in Lacs per annum?)": "Current CTC (LPA)",
    "Notice period/ Availability to join": "Notice Period",
    "Curr. Company name": "Company",
    "Curr. Company Designation": "Designation",
    "Under Graduation degree": "UG Degree",
    "UG University/institute Name": "UG University",
    "PG specialization": "PG Specialization",
    "PG university/institute name": "PG University",
    "Doctorate specialization": "Doctorate Spec.",
    "Doctorate university/institute name": "Doctorate University",
    "Latest Pipeline Stage": "Pipeline Stage",
    "Home Town/City": "Home Town",
    "Date of application": "Applied On",
    "Candidate profile": "Profile Link",
    "Annual Salary": "Salary",
    "Current Location": "Location",
    "Preferred Locations": "Preferred Locations",
    "Total Experience": "Experience",
    "Resume Headline": "Resume Headline",
    "Phone Number": "Phone",
    "Email ID": "Email",
    "Date of Birth": "DOB",
}

# Columns with higher keyword relevance for ATS scoring
ATS_HIGH_WEIGHT_COLS = {"Ans(Key Skills)", "Resume Headline", "Curr. Company Designation"}
ATS_MEDIUM_WEIGHT_COLS = {"Job Title", "Total Experience", "Curr. Company name",
                          "Under Graduation degree", "PG specialization"}

# Default columns to show (user can change)
DEFAULT_VISIBLE_COLS = [
    "Name", "Email ID", "Phone Number", "Current Location", "Total Experience",
    "Curr. Company name", "Curr. Company Designation", "Annual Salary",
    "Notice period/ Availability to join", "Resume Headline", "Ans(Key Skills)",
    "Latest Pipeline Stage", "Candidate profile",
]


# ── Dark Theme CSS ───────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    .stApp, .main, .block-container { background-color: #0e0e1a !important; }
    section[data-testid="stSidebar"] { background-color: #12121f !important; }
    section[data-testid="stSidebar"] > div { background-color: #12121f !important; }
    #MainMenu, footer { visibility: hidden; }
    header[data-testid="stHeader"] { background: transparent !important; }

    .app-header {
        background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 1.5rem;
        border: 1px solid rgba(139,92,246,0.15);
        box-shadow: 0 8px 32px rgba(0,0,0,0.4);
    }
    .app-header h1 { color: #fff; font-size: 1.85rem; font-weight: 700; margin: 0 0 0.3rem 0; }
    .app-header p { color: rgba(255,255,255,0.5); font-size: 0.92rem; margin: 0; }

    .stat-row { display: flex; gap: 1rem; margin-bottom: 1.5rem; flex-wrap: wrap; }
    .stat-card {
        flex: 1; min-width: 120px;
        background: rgba(139,92,246,0.08);
        border: 1px solid rgba(139,92,246,0.18);
        border-radius: 12px;
        padding: 1rem 1.2rem;
        text-align: center;
    }
    .stat-card .stat-value { font-size: 1.7rem; font-weight: 700; color: #a78bfa; line-height: 1.1; }
    .stat-card .stat-label {
        font-size: 0.72rem; color: rgba(255,255,255,0.4);
        text-transform: uppercase; letter-spacing: 0.8px; margin-top: 0.25rem;
    }

    .candidate-card {
        background: #161625;
        border: 1px solid rgba(255,255,255,0.06);
        border-radius: 14px;
        padding: 1.5rem 1.8rem;
        margin-bottom: 1rem;
        transition: border-color 0.25s, box-shadow 0.25s, transform 0.15s;
    }
    .candidate-card:hover {
        border-color: rgba(139,92,246,0.35);
        box-shadow: 0 6px 24px rgba(139,92,246,0.08);
        transform: translateY(-1px);
    }
    .candidate-name { font-size: 1.15rem; font-weight: 600; color: #e0e7ff; margin-bottom: 0.5rem; }
    .candidate-meta {
        display: flex; flex-wrap: wrap; gap: 0.4rem 1.4rem;
        font-size: 0.84rem; color: rgba(255,255,255,0.55); margin-bottom: 0.7rem;
    }
    .candidate-meta span { white-space: nowrap; }
    .candidate-headline {
        font-size: 0.86rem; color: rgba(255,255,255,0.5); line-height: 1.55;
        margin: 0.5rem 0; padding: 0.6rem 0.9rem;
        background: rgba(255,255,255,0.03); border-radius: 8px;
        border-left: 3px solid rgba(139,92,246,0.4);
    }

    .match-pill {
        display: inline-block; background: rgba(139,92,246,0.18); color: #a78bfa;
        padding: 0.2rem 0.6rem; border-radius: 20px; font-size: 0.73rem; font-weight: 500;
        margin: 0.1rem 0.15rem; border: 1px solid rgba(139,92,246,0.25);
    }
    .field-pill {
        display: inline-block; background: rgba(16,185,129,0.12); color: #6ee7b7;
        padding: 0.2rem 0.6rem; border-radius: 20px; font-size: 0.73rem; font-weight: 500;
        margin: 0.1rem 0.15rem; border: 1px solid rgba(16,185,129,0.2);
    }

    .profile-link {
        display: inline-block;
        background: linear-gradient(135deg, #6366f1, #8b5cf6);
        color: #fff !important; text-decoration: none !important;
        padding: 0.4rem 1.1rem; border-radius: 8px;
        font-size: 0.8rem; font-weight: 500; margin-top: 0.5rem;
        transition: opacity 0.2s, transform 0.15s;
    }
    .profile-link:hover { opacity: 0.85; transform: translateY(-1px); }

    .detail-grid {
        display: grid; grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
        gap: 0.35rem 1.4rem; font-size: 0.82rem; color: rgba(255,255,255,0.5); margin: 0.5rem 0;
    }
    .detail-item { display: flex; gap: 0.4rem; }
    .detail-label { color: rgba(255,255,255,0.3); min-width: 110px; flex-shrink: 0; }
    .detail-value { color: rgba(255,255,255,0.65); word-break: break-word; }

    .sidebar-section {
        color: #a78bfa; font-size: 0.72rem; font-weight: 600;
        text-transform: uppercase; letter-spacing: 1.2px;
        margin: 1.2rem 0 0.5rem 0; padding-bottom: 0.35rem;
        border-bottom: 1px solid rgba(139,92,246,0.15);
    }

    .stTextInput input, .stSelectbox select, .stMultiSelect {
        background-color: #1a1a2e !important; color: #e0e0f0 !important;
    }
    div[data-testid="stExpander"] {
        background-color: #161625 !important;
        border-color: rgba(255,255,255,0.06) !important;
        border-radius: 12px !important;
    }

    .no-results {
        text-align: center; padding: 3rem 2rem; background: #161625;
        border: 1px dashed rgba(255,255,255,0.1); border-radius: 14px; margin-top: 1rem;
    }
    .no-results h3 { color: rgba(255,255,255,0.5); font-weight: 500; }
    .no-results p { color: rgba(255,255,255,0.3); font-size: 0.9rem; }

    /* ATS Score badge */
    .ats-badge {
        display: inline-flex; align-items: center; justify-content: center;
        width: 52px; height: 52px; border-radius: 50%; font-size: 1rem; font-weight: 700;
        margin-right: 1rem; flex-shrink: 0;
    }
    .ats-high { background: rgba(16,185,129,0.2); color: #6ee7b7; border: 2px solid rgba(16,185,129,0.4); }
    .ats-mid  { background: rgba(251,191,36,0.2); color: #fbbf24; border: 2px solid rgba(251,191,36,0.4); }
    .ats-low  { background: rgba(239,68,68,0.15); color: #f87171; border: 2px solid rgba(239,68,68,0.3); }

    .ats-row {
        display: flex; align-items: center; margin-bottom: 0.6rem;
    }
</style>
""", unsafe_allow_html=True)


# ── Data helpers ─────────────────────────────────────────────────────────────

def _parse_ctc(val):
    if val is None or str(val).strip() in ("N/A", "NA", ""):
        return 0.0
    try:
        v = float(val)
        return round(v / 100000, 2) if v > 10000 else v
    except (ValueError, TypeError):
        return 0.0


def _cell_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _extract_candidates(ws, headers):
    """Extract all candidate rows with all columns, plus computed fields."""
    total = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=3).value:  # Name column
            total += 1
        else:
            break

    candidates = []
    for row in range(2, 2 + total):
        c = {}
        for col_idx, header in headers.items():
            val = ws.cell(row=row, column=col_idx).value
            c[header] = _cell_str(val)

        # Profile hyperlink (actual URL, not display text)
        if "Candidate profile" in headers.values():
            profile_col = [k for k, v in headers.items() if v == "Candidate profile"][0]
            cell = ws.cell(row=row, column=profile_col)
            if cell.hyperlink:
                c["_profile_url"] = cell.hyperlink.target
            else:
                c["_profile_url"] = c.get("Candidate profile", "")

        # Computed fields for filtering
        exp_str = c.get("Total Experience", "")
        m = re.search(r"(\d+)\s*Year", exp_str, re.IGNORECASE)
        c["_exp_years"] = int(m.group(1)) if m else 0

        c["_current_ctc"] = _parse_ctc(
            c.get("Ans(What is your current CTC in Lacs per annum?)",
                  c.get("Annual Salary", ""))
        )
        c["_expected_ctc"] = _parse_ctc(
            c.get("Ans(What is your expected CTC in Lacs per annum?)", "")
        )

        candidates.append(c)

    return candidates


def _get_headers(ws):
    """Read column headers from row 1."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value).strip()
    return headers


@st.cache_data
def load_local_file(filepath, sheet_name):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    headers = _get_headers(ws)
    return headers, _extract_candidates(ws, headers)


@st.cache_data
def load_uploaded_file(file_bytes, sheet_name):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[sheet_name]
    headers = _get_headers(ws)
    return headers, _extract_candidates(ws, headers)


@st.cache_data
def get_sheet_names(filepath):
    return openpyxl.load_workbook(filepath, read_only=True).sheetnames


@st.cache_data
def get_sheet_names_bytes(file_bytes):
    return openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True).sheetnames


def short_name(header):
    """Get a shorter display name for a column header."""
    return HEADER_SHORT_NAMES.get(header, header)


# ── ATS Score ────────────────────────────────────────────────────────────────

def compute_ats_score(candidate, keywords, all_headers):
    """
    Compute an ATS (Applicant Tracking System) match score (0-100).

    Algorithm:
    - For each keyword, search across all text fields in the candidate record
    - High-weight fields (Key Skills, Resume Headline, Designation): 3x
    - Medium-weight fields (Job Title, Experience, Company, Education): 2x
    - All other text fields: 1x
    - Score = (weighted_hits / max_possible_weighted_score) * 100
    - Bonus for keyword density (multiple occurrences in important fields)
    """
    if not keywords:
        return 0

    max_score = 0
    earned_score = 0

    for kw in keywords:
        for header in all_headers:
            val = candidate.get(header, "").lower()
            if not val or val in ("n/a", "na", ""):
                continue

            if header in ATS_HIGH_WEIGHT_COLS:
                weight = 3.0
            elif header in ATS_MEDIUM_WEIGHT_COLS:
                weight = 2.0
            else:
                weight = 1.0

            max_score += weight

            if kw in val:
                earned_score += weight
                # Density bonus: extra 0.5 for each additional occurrence
                count = val.count(kw)
                if count > 1:
                    earned_score += min((count - 1) * 0.3, 1.0)

        # Check fetched resume text (highest weight - actual resume content)
        fetched_resume = candidate.get("_fetched_resume", "")
        if fetched_resume:
            resume_lower = fetched_resume.lower()
            resume_weight = 4.0
            max_score += resume_weight
            if kw in resume_lower:
                earned_score += resume_weight
                count = resume_lower.count(kw)
                if count > 1:
                    earned_score += min((count - 1) * 0.2, 2.0)

    if max_score == 0:
        return 0

    raw = (earned_score / max_score) * 100
    # Scale up: since most keywords won't appear in every field,
    # normalize so that matching in key fields gives a good score
    scaled = min(raw * 3.5, 100)
    return round(scaled)


def ats_badge_class(score):
    if score >= 60:
        return "ats-high"
    elif score >= 30:
        return "ats-mid"
    return "ats-low"


# ── Search ───────────────────────────────────────────────────────────────────

def search_candidates(candidates, keywords, match_mode, search_in_headers):
    """Search candidates, return matches with match metadata."""
    if not keywords:
        return []

    results = []
    for c in candidates:
        combined_texts = {}
        for header in search_in_headers:
            text = c.get(header, "").lower()
            if text and text not in ("n/a", "na"):
                combined_texts[header] = text

        full_text = " ".join(combined_texts.values())
        matched_kw = []
        matched_fields = []

        for kw in keywords:
            if kw in full_text:
                matched_kw.append(kw)
                fields = [short_name(h) for h, t in combined_texts.items() if kw in t]
                matched_fields.extend(fields)

        if match_mode == "All Keywords" and len(matched_kw) < len(keywords):
            continue
        if match_mode == "Any Keyword" and not matched_kw:
            continue

        result = dict(c)
        result["_matched_kw"] = list(set(matched_kw))
        result["_matched_fields"] = list(set(matched_fields))
        results.append(result)

    return results


def apply_filters(results, exp_range, ctc_range, ctc_type):
    filtered = []
    for c in results:
        exp = c.get("_exp_years", 0)
        if exp < exp_range[0] or exp > exp_range[1]:
            continue
        ctc = c.get("_current_ctc", 0) if ctc_type == "Current CTC" else c.get("_expected_ctc", 0)
        if ctc < ctc_range[0] or ctc > ctc_range[1]:
            continue
        filtered.append(c)
    return filtered


def results_to_excel(results, visible_cols, keywords, all_headers):
    """Export results to styled Excel bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Results"

    export_cols = ["#", "ATS Score"] + [short_name(h) for h in visible_cols] + \
                  ["Matched Keywords", "Matched Fields", "Profile URL"]

    for col, h in enumerate(export_cols, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F46E5")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    for i, c in enumerate(results, 1):
        ats = compute_ats_score(c, keywords, all_headers)
        row_data = [i, ats] + [c.get(h, "") for h in visible_cols] + [
            ", ".join(c.get("_matched_kw", [])),
            ", ".join(c.get("_matched_fields", [])),
            c.get("_profile_url", ""),
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i + 1, column=col, value=val)

    for col in range(1, len(export_cols) + 1):
        mx = max(len(str(ws.cell(row=r, column=col).value or ""))
                 for r in range(1, len(results) + 2))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Naukri Integration ───────────────────────────────────────────────────────

def naukri_login(email, password):
    """
    Login to Naukri recruiter portal using Selenium (headless Chrome).
    Returns (success: bool, cookies: dict, message: str).
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1280,900")
        opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        driver = webdriver.Chrome(options=opts)
        driver.set_page_load_timeout(30)

        try:
            driver.get("https://www.naukri.com/recruit/login")
            time.sleep(3)

            # Wait for email field to appear
            wait = WebDriverWait(driver, 15)

            # Try multiple selectors for email input
            email_sel = None
            for sel in ['input[type="text"]', 'input[name="email"]', 'input[placeholder*="email"]',
                        'input[placeholder*="Email"]', '#email', '#usernameField']:
                try:
                    email_sel = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
                    break
                except:
                    continue

            if not email_sel:
                driver.quit()
                return False, None, "Could not find email field on Naukri login page."

            email_sel.clear()
            email_sel.send_keys(email)
            time.sleep(0.5)

            # Find password field
            pass_sel = None
            for sel in ['input[type="password"]', 'input[name="password"]', '#passwordField']:
                try:
                    pass_sel = driver.find_element(By.CSS_SELECTOR, sel)
                    break
                except:
                    continue

            if not pass_sel:
                driver.quit()
                return False, None, "Could not find password field on Naukri login page."

            pass_sel.clear()
            pass_sel.send_keys(password)
            time.sleep(0.5)

            # Find and click login button
            btn = None
            for sel in ['button[type="submit"]', 'input[type="submit"]',
                        'button:not([disabled])', '.btn-primary']:
                try:
                    btn = driver.find_element(By.CSS_SELECTOR, sel)
                    break
                except:
                    continue

            if btn:
                btn.click()
            else:
                from selenium.webdriver.common.keys import Keys
                pass_sel.send_keys(Keys.RETURN)

            time.sleep(5)

            # Check if login was successful by looking at the URL or page content
            current_url = driver.current_url
            if "login" in current_url.lower() and "dashboard" not in current_url.lower():
                # Might still be on login page — check for error messages
                page_text = driver.page_source.lower()
                if "invalid" in page_text or "incorrect" in page_text or "wrong" in page_text:
                    driver.quit()
                    return False, None, "Login failed — invalid email or password."
                # Give more time for redirect
                time.sleep(3)
                current_url = driver.current_url

            # Extract cookies for requests session
            selenium_cookies = driver.get_cookies()
            cookies = {}
            for c in selenium_cookies:
                cookies[c["name"]] = c["value"]

            driver.quit()

            if cookies:
                return True, cookies, "Successfully connected to Naukri!"
            else:
                return False, None, "Login completed but no session cookies received."

        except Exception as e:
            driver.quit()
            return False, None, f"Login error: {str(e)}"

    except ImportError:
        return False, None, "Selenium is not installed. Run: pip install selenium"
    except Exception as e:
        return False, None, f"Browser error: {str(e)}"


def fetch_naukri_profile(profile_url, cookies):
    """
    Fetch a candidate profile page from Naukri using session cookies.
    Returns the extracted text content from the profile page.
    """
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://hiring.naukri.com/",
        }
        r = requests.get(profile_url, cookies=cookies, headers=headers,
                         timeout=15, allow_redirects=True)

        if r.status_code != 200:
            return None, f"HTTP {r.status_code}"

        # If redirected to login, cookies expired
        if "recruit/login" in r.url:
            return None, "Session expired — please reconnect."

        soup = BeautifulSoup(r.text, "html.parser")

        # Remove script/style tags
        for tag in soup(["script", "style", "noscript", "nav", "footer", "header"]):
            tag.decompose()

        # Get all visible text
        text = soup.get_text(separator=" ", strip=True)

        # Clean up whitespace
        text = re.sub(r"\s+", " ", text).strip()

        return text, "OK"

    except requests.Timeout:
        return None, "Request timed out"
    except Exception as e:
        return None, str(e)


def batch_fetch_profiles(results, cookies, progress_callback=None):
    """Fetch profile data for all results with profile URLs."""
    fetched = {}
    total = len(results)
    for idx, c in enumerate(results):
        url = c.get("_profile_url", "")
        if url and url.startswith("http"):
            if url in st.session_state.naukri_fetched:
                fetched[url] = st.session_state.naukri_fetched[url]
            else:
                text, status = fetch_naukri_profile(url, cookies)
                fetched[url] = {"text": text, "status": status}
                st.session_state.naukri_fetched[url] = fetched[url]
                time.sleep(0.5)  # Rate limiting

        if progress_callback:
            progress_callback((idx + 1) / total)

    return fetched


# ══════════════════════════════════════════════════════════════════════════════
# ── App Header ───────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
    <h1>✅ ATS Checker</h1>
    <p>Search candidate databases by skills, experience, location, CTC — with ATS resume scoring.</p>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ── Sidebar ──────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:

    # ── Data Source ──
    st.markdown('<div class="sidebar-section">📂 Data Source</div>', unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Upload Excel files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Upload .xlsx candidate files. Combined with the default local file if present.",
        label_visibility="collapsed",
    )

    file_sources = {}
    if os.path.exists(EXCEL_FILE):
        file_sources[f"📁 {os.path.basename(EXCEL_FILE)}"] = {"type": "local", "path": EXCEL_FILE}
    if uploaded_files:
        for uf in uploaded_files:
            file_sources[f"📤 {uf.name}"] = {"type": "upload", "bytes": uf.getvalue(), "name": uf.name}

    if not file_sources:
        st.warning("No files available. Upload an Excel file to get started.")
        st.stop()

    selected_files = st.multiselect(
        "Select files to search",
        list(file_sources.keys()),
        default=list(file_sources.keys()),
    )
    if not selected_files:
        st.info("Select at least one file.")
        st.stop()

    # Load all candidates and collect all unique headers
    candidates = []
    all_headers_set = []
    active_file_count = len(selected_files)

    for fname in selected_files:
        src = file_sources[fname]
        if src["type"] == "local":
            sheets = get_sheet_names(src["path"])
        else:
            sheets = get_sheet_names_bytes(src["bytes"])

        sel_sheet = st.selectbox(f"Sheet — {fname}", sheets, index=0, key=f"sh_{fname}")

        if src["type"] == "local":
            headers, file_candidates = load_local_file(src["path"], sel_sheet)
        else:
            headers, file_candidates = load_uploaded_file(src["bytes"], sel_sheet)

        for h in headers.values():
            if h not in all_headers_set:
                all_headers_set.append(h)

        for c in file_candidates:
            c["_source_file"] = fname
        candidates.extend(file_candidates)

    # ── Column Visibility ──
    st.markdown('<div class="sidebar-section">📊 Column Visibility</div>', unsafe_allow_html=True)

    default_vis = [h for h in DEFAULT_VISIBLE_COLS if h in all_headers_set]
    visible_cols = st.multiselect(
        "Columns to display",
        all_headers_set,
        default=default_vis,
        format_func=short_name,
        help="Select which columns to show in candidate cards and preview table.",
    )

    # ── Keywords ──
    st.markdown('<div class="sidebar-section">🏷️ Keywords</div>', unsafe_allow_html=True)

    if "keywords" not in st.session_state:
        st.session_state.keywords = []

    kw_col1, kw_col2 = st.columns([3, 1])
    with kw_col1:
        new_kw = st.text_input("kw", placeholder="Type a keyword...",
                               label_visibility="collapsed", key="kw_input")
    with kw_col2:
        add_clicked = st.button("Add", use_container_width=True, key="add_kw")

    if add_clicked and new_kw.strip():
        for k in new_kw.split(","):
            k = k.strip().lower()
            if k and k not in st.session_state.keywords:
                st.session_state.keywords.append(k)

    if st.session_state.keywords:
        chip_cols = st.columns(min(len(st.session_state.keywords), 4))
        remove_idx = None
        for idx, kw in enumerate(st.session_state.keywords):
            with chip_cols[idx % len(chip_cols)]:
                if st.button(f"✕ {kw}", key=f"rm_{idx}", use_container_width=True):
                    remove_idx = idx
        if remove_idx is not None:
            st.session_state.keywords.pop(remove_idx)
            st.rerun()
        if st.button("Clear all keywords", use_container_width=True):
            st.session_state.keywords = []
            st.rerun()

    match_mode = st.radio("Match Mode", ["Any Keyword", "All Keywords"], index=0, horizontal=True)

    search_in = st.multiselect(
        "Search in fields",
        all_headers_set,
        default=all_headers_set,
        format_func=short_name,
    )

    # ── Experience ──
    st.markdown('<div class="sidebar-section">💼 Experience</div>', unsafe_allow_html=True)

    all_exp = [c["_exp_years"] for c in candidates]
    min_e, max_e = (min(all_exp), max(all_exp)) if all_exp else (0, 20)
    exp_range = st.slider("Experience (years)", 0, max(max_e, 20), (min_e, max_e), 1)

    # ── CTC Range ──
    st.markdown('<div class="sidebar-section">💰 CTC Range</div>', unsafe_allow_html=True)

    ctc_type = st.radio("CTC type", ["Current CTC", "Expected CTC"], index=0, horizontal=True)
    if ctc_type == "Current CTC":
        all_ctc = [c["_current_ctc"] for c in candidates if c["_current_ctc"] > 0]
    else:
        all_ctc = [c["_expected_ctc"] for c in candidates if c["_expected_ctc"] > 0]
    max_ctc = max(all_ctc) if all_ctc else 25.0
    ctc_range = st.slider("CTC (Lakhs/annum)", 0.0, max(max_ctc, 30.0),
                          (0.0, max(max_ctc, 30.0)), 0.5, format="₹%.1fL")

    # ── Sort ──
    st.markdown('<div class="sidebar-section">🔃 Sort</div>', unsafe_allow_html=True)

    sort_by = st.selectbox("Sort by", [
        "ATS Score (High→Low)", "ATS Score (Low→High)",
        "Name", "Experience (Low→High)", "Experience (High→Low)",
        "Current CTC (Low→High)", "Current CTC (High→Low)",
    ], index=0, label_visibility="collapsed")

    # ── Naukri Credentials ──
    st.markdown('<div class="sidebar-section">🔐 Naukri Login</div>', unsafe_allow_html=True)

    if "naukri_logged_in" not in st.session_state:
        st.session_state.naukri_logged_in = False
    if "naukri_cookies" not in st.session_state:
        st.session_state.naukri_cookies = None
    if "naukri_fetched" not in st.session_state:
        st.session_state.naukri_fetched = {}

    if st.session_state.naukri_logged_in:
        st.success("Connected to Naukri")
        if st.button("Disconnect", use_container_width=True):
            st.session_state.naukri_logged_in = False
            st.session_state.naukri_cookies = None
            st.session_state.naukri_fetched = {}
            st.rerun()
    else:
        naukri_email = st.text_input("Naukri Email", placeholder="recruiter@company.com",
                                     key="nk_email", label_visibility="collapsed")
        naukri_pass = st.text_input("Naukri Password", type="password",
                                    placeholder="Password", key="nk_pass",
                                    label_visibility="collapsed")
        if st.button("🔐 Connect to Naukri", use_container_width=True):
            if naukri_email and naukri_pass:
                with st.spinner("Logging in to Naukri..."):
                    ok, cookies, msg = naukri_login(naukri_email, naukri_pass)
                if ok:
                    st.session_state.naukri_logged_in = True
                    st.session_state.naukri_cookies = cookies
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
            else:
                st.warning("Enter both email and password.")

        st.caption("Credentials are used only for this session and never stored.")

    st.markdown("---")
    st.button("🔍  Search Candidates", type="primary", use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# ── Main Content ─────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

keywords = st.session_state.keywords

# Stats
st.markdown(f"""
<div class="stat-row">
    <div class="stat-card">
        <div class="stat-value">{len(candidates)}</div>
        <div class="stat-label">Total Candidates</div>
    </div>
    <div class="stat-card">
        <div class="stat-value">{active_file_count}</div>
        <div class="stat-label">Files Loaded</div>
    </div>
    <div class="stat-card">
        <div class="stat-value">{len(keywords)}</div>
        <div class="stat-label">Keywords</div>
    </div>
    <div class="stat-card">
        <div class="stat-value">{exp_range[0]}–{exp_range[1]} yrs</div>
        <div class="stat-label">Exp. Filter</div>
    </div>
    <div class="stat-card">
        <div class="stat-value">₹{ctc_range[0]:.0f}–{ctc_range[1]:.0f}L</div>
        <div class="stat-label">{ctc_type} Filter</div>
    </div>
</div>
""", unsafe_allow_html=True)

if keywords:
    results = search_candidates(candidates, keywords, match_mode, search_in)
    results = apply_filters(results, exp_range, ctc_range, ctc_type)

    # Compute ATS scores
    for c in results:
        c["_ats_score"] = compute_ats_score(c, keywords, all_headers_set)

    # Sort
    if sort_by == "ATS Score (High→Low)":
        results.sort(key=lambda c: c.get("_ats_score", 0), reverse=True)
    elif sort_by == "ATS Score (Low→High)":
        results.sort(key=lambda c: c.get("_ats_score", 0))
    elif sort_by == "Experience (Low→High)":
        results.sort(key=lambda c: c.get("_exp_years", 0))
    elif sort_by == "Experience (High→Low)":
        results.sort(key=lambda c: c.get("_exp_years", 0), reverse=True)
    elif sort_by == "Current CTC (Low→High)":
        results.sort(key=lambda c: c.get("_current_ctc", 0))
    elif sort_by == "Current CTC (High→Low)":
        results.sort(key=lambda c: c.get("_current_ctc", 0), reverse=True)
    else:
        results.sort(key=lambda c: str(c.get("Name", "")).lower())

    # Results summary
    avg_ats = round(sum(c["_ats_score"] for c in results) / len(results)) if results else 0
    st.markdown(f"""
    <div class="stat-row">
        <div class="stat-card">
            <div class="stat-value">{len(results)}</div>
            <div class="stat-label">Matches Found</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{avg_ats}%</div>
            <div class="stat-label">Avg ATS Score</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{match_mode.split()[0]}</div>
            <div class="stat-label">Match Mode</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if results:
        # Action buttons row
        btn_col1, btn_col2, btn_col3 = st.columns([2, 1, 1])
        with btn_col2:
            if st.session_state.naukri_logged_in:
                fetch_clicked = st.button("🌐 Fetch Resumes from Naukri", use_container_width=True)
            else:
                st.button("🌐 Fetch Resumes (login first)", disabled=True, use_container_width=True)
                fetch_clicked = False
        with btn_col3:
            excel_bytes = results_to_excel(results, visible_cols, keywords, all_headers_set)
            st.download_button(
                label="📥 Export to Excel",
                data=excel_bytes,
                file_name=f"search_{'_'.join(keywords)[:40]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # Fetch profiles if requested
        if fetch_clicked and st.session_state.naukri_logged_in:
            progress = st.progress(0, text="Fetching profiles from Naukri...")
            fetched = batch_fetch_profiles(
                results, st.session_state.naukri_cookies,
                progress_callback=lambda p: progress.progress(p, text=f"Fetching profiles... {int(p*100)}%")
            )
            progress.empty()

            # Recompute ATS scores with fetched resume text
            for c in results:
                url = c.get("_profile_url", "")
                if url and url in fetched and fetched[url]["text"]:
                    c["_fetched_resume"] = fetched[url]["text"]
                    c["_ats_score"] = compute_ats_score(c, keywords, all_headers_set)

            # Re-sort after score update
            if "ATS Score" in sort_by:
                reverse = "High" in sort_by
                results.sort(key=lambda c: c.get("_ats_score", 0), reverse=reverse)

            st.success(f"Fetched {sum(1 for v in fetched.values() if v['text'])} / {len(fetched)} profiles successfully.")

        # Enhance ATS with any previously fetched data
        for c in results:
            url = c.get("_profile_url", "")
            if url and url in st.session_state.naukri_fetched:
                cached = st.session_state.naukri_fetched[url]
                if cached.get("text"):
                    c["_fetched_resume"] = cached["text"]
                    c["_ats_score"] = compute_ats_score(c, keywords, all_headers_set)

        # Candidate cards
        def esc(val):
            """HTML-escape a value for safe embedding."""
            if val is None:
                return ""
            return html_mod.escape(str(val))

        for i, c in enumerate(results, 1):
            ats = c["_ats_score"]
            badge_cls = ats_badge_class(ats)
            kw_pills = "".join(f'<span class="match-pill">{esc(kw)}</span>' for kw in c.get("_matched_kw", []))
            field_pills = "".join(f'<span class="field-pill">{esc(f)}</span>' for f in c.get("_matched_fields", []))

            profile_url = c.get("_profile_url", "")
            profile_html = ""
            if profile_url and str(profile_url).startswith("http"):
                profile_html = f'<a class="profile-link" href="{esc(profile_url)}" target="_blank">View Profile &#8594;</a>'
                # Show fetch status
                if profile_url in st.session_state.naukri_fetched:
                    fetch_info = st.session_state.naukri_fetched[profile_url]
                    if fetch_info.get("text"):
                        profile_html += ' <span style="color:#6ee7b7;font-size:0.75rem;">&#10003; Resume fetched</span>'
                    else:
                        profile_html += f' <span style="color:rgba(255,255,255,0.3);font-size:0.75rem;">&#9888; {esc(fetch_info.get("status",""))}</span>'

            # Build detail grid from visible columns (exclude Name which is in the header)
            detail_items = []
            for h in visible_cols:
                if h in ("Name", "Candidate profile"):
                    continue
                val = c.get(h, "")
                if val and str(val).strip() not in ("N/A", "NA", ""):
                    detail_items.append(
                        '<div class="detail-item">'
                        f'<span class="detail-label">{esc(short_name(h))}</span>'
                        f'<span class="detail-value">{esc(val)}</span>'
                        '</div>'
                    )
            detail_grid_html = "\n".join(detail_items)

            # Resume headline (always shown if available)
            headline = c.get("Resume Headline", "")
            headline_section = ""
            if headline and str(headline).strip() not in ("N/A", "NA", ""):
                headline_section = f'<div class="candidate-headline">{esc(headline)}</div>'

            # Key skills
            skills = c.get("Ans(Key Skills)", "")
            skills_section = ""
            if skills and str(skills).strip() not in ("N/A", "NA", ""):
                skills_section = f'<div style="font-size:0.82rem;color:rgba(255,255,255,0.4);margin-top:0.3rem;"><strong style="color:rgba(255,255,255,0.5);">Key Skills:</strong> {esc(skills)}</div>'

            # Source file tag
            source_tag = ""
            if active_file_count > 1:
                source_tag = f'<div style="font-size:0.72rem;color:rgba(139,92,246,0.6);margin-bottom:0.3rem;">{esc(c.get("_source_file",""))}</div>'

            # Quick meta line
            name_val = esc(c.get("Name", ""))
            email_val = esc(c.get("Email ID", ""))
            phone_val = esc(c.get("Phone Number", ""))
            loc_val = esc(c.get("Current Location", ""))
            exp_val = esc(c.get("Total Experience", ""))

            card_html = (
                '<div class="candidate-card">'
                '<div class="ats-row">'
                f'<div class="ats-badge {badge_cls}">{ats}%</div>'
                '<div>'
                f'<div class="candidate-name">#{i} &nbsp; {name_val}</div>'
                f'{source_tag}'
                '</div>'
                '</div>'
                '<div class="candidate-meta">'
                f'<span>&#9993; {email_val}</span>'
                f'<span>&#9742; {phone_val}</span>'
                f'<span>&#128205; {loc_val}</span>'
                f'<span>&#128188; {exp_val}</span>'
                '</div>'
                f'<div class="detail-grid">{detail_grid_html}</div>'
                f'{headline_section}'
                f'{skills_section}'
                '<div style="margin-top:0.6rem;">'
                f'<span style="font-size:0.75rem;color:rgba(255,255,255,0.35);">Matched:</span> {kw_pills}'
                f'<span style="font-size:0.75rem;color:rgba(255,255,255,0.35);margin-left:0.6rem;">In:</span> {field_pills}'
                '</div>'
                f'<div style="margin-top:0.6rem;">{profile_html}</div>'
                '</div>'
            )
            st.markdown(card_html, unsafe_allow_html=True)

    else:
        st.markdown("""
        <div class="no-results">
            <h3>No candidates matched</h3>
            <p>Try broader keywords, switch to "Any Keyword" mode, or widen the experience / CTC filters.</p>
        </div>
        """, unsafe_allow_html=True)

else:
    # Landing — full preview with all columns
    st.info("👆 Add keywords in the sidebar to search. Use **Column Visibility** to choose which columns to display.")

    with st.expander("📊 Preview all candidates (all columns)", expanded=False):
        preview = []
        display_headers = visible_cols if visible_cols else all_headers_set
        for c in candidates:
            row = {}
            for h in display_headers:
                row[short_name(h)] = c.get(h, "")
            preview.append(row)
        st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)

    st.markdown("""
    > **📝 ATS Score Note:** The ATS (Applicant Tracking System) score is computed from the candidate data
    > available in the Excel file — Resume Headline, Key Skills, Experience, Education, Designation, and all
    > other fields. Profile links point to Naukri recruiter portal which requires login credentials for access.
    > The score reflects how well the candidate's listed qualifications match your search keywords.
    """)
